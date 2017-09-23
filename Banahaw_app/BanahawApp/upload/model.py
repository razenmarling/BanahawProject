"""."""

import os
import openpyxl
from sqlalchemy import desc
from BanahawApp import Session
from BanahawApp.table import T_Member00, T_Member01


class Uploadmodel(object):
	"""."""
	def __init__(self, filepath, filename):
		"""."""
		self.__session = Session()
		self.__filepath = filepath
		self.__filename = filename
		self.__dict_data = {}

	def process_file(self):
		"""'Personalized', 'Family', 'Upgraded'."""
		funcs = {
			'Personalized': self.__process_personalized,
			'Family': self.__process_family,
			'Upgraded': self.__process_upgraded
		}

		if os.path.exists(self.__filepath):
			sheets = self.__open_xlsx()

			for s in sheets:
				active_sheet = self.__wb.get_sheet_by_name(s)
				funcs[s](active_sheet)

	def __open_xlsx(self):
		"""."""
		self.__wb=openpyxl.load_workbook(self.__filepath)

		return self.__wb.get_sheet_names()

	def __process_personalized(self, sheet):
		"""."""
		data_list = self.__extract_data(sheet)

		for data_dict in data_list:
			member = T_Member00()

			data_dict['membertype'] = 'Personalized'
			data_dict['membershipcost'] = 300
			data_dict['attendant_name'] = 'Admin'
			data_dict['attendantid'] = 0
			data_dict['branch'] = 'Plaridel'

			for key, value in data_dict.items():

				if key == 'DATE_APPLIED':
					key = 'datecreated'

				setattr(member, key.lower(), value)

			self.__session.add(member)

		self.__session.commit()

	def __process_family(self, sheet):
		"""."""
		submem_param = []
		data_list = self.__extract_data(sheet)

		for data_dict in data_list:
			member = T_Member00()

			data_dict['membertype'] = 'Family'
			data_dict['membershipcost'] = 600
			data_dict['attendant_name'] = 'Admin'
			data_dict['attendantid'] = 0
			data_dict['branch'] = 'Plaridel'

			temp_dict = {}
			for key, value in data_dict.items():

				if key in ['SUBMEMBERS', 'BIRTHDATE', 'NAME', 'DATE_APPLIED']:

					if key == 'DATE_APPLIED':
						key = 'datecreated'

					temp_dict[key.lower()] = value

					if key == 'SUBMEMBERS':
						continue

				setattr(member, key.lower(), value)

			self.__session.add(member)
			submem_param.append(temp_dict)

		self.__session.commit()

		self.__insert_submem(submem_param)

	def __process_upgraded(self, sheet):
		"""."""
		search_cols = ['NAME', 'BIRTHDATE']
		mem01_search_cols = ['SUBMEMBERS', 'BIRTHDATE', 'NAME', 'DATE_UPGRADED']
		data_list = self.__extract_data(sheet)
		mem01_search_param = []

		for data_dict in data_list:
			search_param = []
			temp_dict = {}
			member = T_Member00()

			data_dict['membertype'] = 'Family'
			data_dict['membershipcost'] = 600
			data_dict['attendant_name'] = 'Admin'
			data_dict['attendantid'] = '0'
			data_dict['branch'] = 'Plaridel'

			for key, value in data_dict.items():
				if key in search_cols and value:
					search_param.append(getattr(T_Member00, key.lower())==value)

			result = self.__session.query(T_Member00).filter(*search_param).order_by(
					 desc(T_Member00.datecreated)).first()

			if result:
				# update
				update_param = ['membertype', 'membershipcost', 'attendant_name',
								'attendantid', 'branch']

				for key, value in data_dict.items():

					if key in update_param and value:
						setattr(result, key.lower(), value)

					if key in mem01_search_cols:

						if key == 'DATE_UPGRADED':
							key = 'datecreated'

						temp_dict[key.lower()] = value

				mem01_search_param.append(temp_dict)

			else:
				# insert
				for key, value in data_dict.items():

					if key in mem01_search_cols:

						if key == 'DATE_UPGRADED':
							key = 'datecreated'

						temp_dict[key.lower()] = value

						if key == 'SUBMEMBERS':
							continue

					setattr(member, key.lower(), value)

				self.__session.add(member)
				mem01_search_param.append(temp_dict)

		self.__session.commit()

		self.__insert_submem(mem01_search_param)

	def __extract_data(self, sheet):
		"""."""
		headers = []
		retval = []
		row = sheet.max_row + 1
		col = sheet.max_column + 1

		for row_num in range(1, row):
			temp_dict = {}

			for col_num in range(1, col):
				if row_num == 1:
					headers.append(sheet.cell(row=row_num, column=col_num).value)
				else:
					key = headers[col_num - 1]
					value = sheet.cell(row=row_num, column=col_num).value
					temp_dict[key] = '' if not value else value

			if temp_dict:
				retval.append(temp_dict)

		return retval

	def __insert_submem(self, param):
		"""."""
		memid = 0
		for data in param:
			print(data)
			search_param = []
			for key, value in data.items():
				if key == 'submembers':
					continue

				search_param.append(getattr(T_Member00, key) == value)

			result = self.__session.query(T_Member00).filter(*search_param).order_by(
					 desc(T_Member00.datecreated)).first()

			if result:
				print('xxx')
				memid = result.member00id

				for sub_mem in data['submembers'].split(','):

					if not sub_mem:
						continue

					member01 = T_Member01()

					setattr(member01, 'member00id', memid)
					setattr(member01, 'relationship', 'Family')
					setattr(member01, 'datecreated', data['datecreated'])
					setattr(member01, 'name', sub_mem)

					self.__session.add(member01)

				self.__session.commit()
